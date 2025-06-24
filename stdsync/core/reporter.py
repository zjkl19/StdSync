# ==================== stdsync/core/reporter.py =============
"""
reporter.py
生成差异报表，并可在原国家公告文件中追加“匹配状态”列
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import List

import xlsxwriter
import openpyxl

from .models import MatchResult

# 颜色映射（状态 → 填充色）
COLOR_MAP = {
    "OBSOLETE": "#FFC7CE",  # 红
    "REVIEW":   "#FFEB9C",  # 黄 / 橙
    "OK":       "#C6EFCE",  # 绿
    "UNUSED":   "#D9D9D9",  # 灰
}

# 差异表列头
HEADERS = [
    "公司标准编号",
    "公司标准名称",
    "国家旧标准号",
    "国家新标准号",
    "实施日期",
    "状态",
    "相似度",
    "查找过程",
]


# ----------------------------------------------------------------------
# 差 异 表 生 成
# ----------------------------------------------------------------------
def render(results: List[MatchResult], out_path: Path | str) -> Path:
    """
    将 MatchResult 列表写入 Excel，并按状态着色

    参数
    ----
    results : List[MatchResult]
        比对结果
    out_path : Path | str
        输出文件路径

    返回
    ----
    Path : 最终生成文件路径
    """
    out_path = Path(out_path)

    wb = xlsxwriter.Workbook(out_path)
    ws = wb.add_worksheet("差异表")

    # 表头格式
    hdr_fmt = wb.add_format({"bold": True, "bg_color": "#B7DEE8"})

    # 写表头
    for col, h in enumerate(HEADERS):
        ws.write(0, col, h, hdr_fmt)

    # 写数据
    for r, m in enumerate(results, start=1):
        ws.write_row(
            r,
            0,
            [
                m.company.code,
                m.company.name,
                m.gb_old_code,
                m.gb_new_code,
                "" if m.company.impl_date is None else str(m.company.impl_date),
                m.status,
                m.similarity,
                m.reason,
            ],
        )

    # 条件格式着色
    max_row = len(results)
    for status, color in COLOR_MAP.items():
        ws.conditional_format(
            1,
            5,
            max_row,
            5,
            {
                "type": "cell",
                "criteria": "==",
                "value": f'"{status}"',
                "format": wb.add_format({"bg_color": color}),
            },
        )

    wb.close()
    return out_path


# ----------------------------------------------------------------------
# 在 国 家 公 告 文 件 追 加 “匹 配 过 程” 列
# ----------------------------------------------------------------------
def append_trace(gb_path: Path | str, results: List[MatchResult]) -> None:
    """
    在国家公告原文件末尾追加一列“匹配状态”，并保存为 *_trace.xlsx

    - 未匹配到公司标准 → 写入 “未使用” 状态
    - 匹配到 → 写入对应状态（OBSOLETE / REVIEW / OK）

    参数
    ----
    gb_path : Path | str
        原国家公告 Excel 路径
    results : List[MatchResult]
        比对结果，用于获取公司标准状态
    """
    gb_path = Path(gb_path)
    wb = openpyxl.load_workbook(gb_path)
    ws = wb.active

    # 追加列标题
    col_idx = ws.max_column + 1
    ws.cell(row=1, column=col_idx).value = "匹配状态"

    # 构建 “公司标准编号 → 状态” 字典
    status_map = {m.company.code: m.status for m in results}

    # 填写每一行状态
    for r in range(2, ws.max_row + 1):
        std_code = ws.cell(row=r, column=1).value  # 默认第一列为编号
        ws.cell(row=r, column=col_idx).value = status_map.get(std_code, "未使用")

    # 另存为 *_trace.xlsx
    out_path = gb_path.with_stem(gb_path.stem + "_trace")
    wb.save(out_path)